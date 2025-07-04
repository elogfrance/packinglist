import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from io import BytesIO
import tempfile

def run():
    st.title("Générateur Packing List Autodoc")

    f1 = st.file_uploader("📁 Fichier F1 (TO SHIP)", type=["xlsx"], key="f1")
    f2 = st.file_uploader("📁 Fichier F2 (E LOG)", type=["xlsx"], key="f2")

    if f1 and f2:
        try:
            # Chargement des fichiers en temporaires
            temp_f1 = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_f1.write(f1.read()); temp_f1.seek(0)
            temp_f2 = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_f2.write(f2.read()); temp_f2.seek(0)

            # Lecture en DataFrame
            df_f1 = pd.read_excel(temp_f1.name)
            df_f2 = pd.read_excel(temp_f2.name)

            # Nettoyage des noms de colonnes
            df_f2.columns = df_f2.columns.str.strip()
            palette_col = next((col for col in df_f2.columns if re.search(r"\bpal\b", col, re.IGNORECASE)), None)
            if not palette_col:
                st.error("❌ Erreur : aucune colonne contenant le mot 'pal' n’a été trouvée dans F2.")
                st.write("Colonnes disponibles :", df_f2.columns.tolist())
                return

            # === Conserver le format d'origine de la colonne palette (colonne E) ===
            df_f2[palette_col] = df_f2[palette_col].astype(str).str.strip()
            # Optionnel : uniformiser la longueur avec zfill
            max_len = df_f2[palette_col].str.len().max()
            df_f2[palette_col] = df_f2[palette_col].str.zfill(max_len)

            # Préparation des colonnes de correspondance
            df_f1["N° COLIS"] = df_f1["Document number"].astype(str).str.strip()
            df_f2["Package Number"] = df_f2["Package Number"].astype(str).str.strip()
            colis_to_palette = dict(zip(df_f2["Package Number"], df_f2[palette_col]))

            # Ordre final des colonnes
            final_order = [
                ("Fournisseur", None),
                ("N° PALETTE", None),
                ("Document number", "N° COLIS"),
                ("Customer order number", "N° groupage"),
                ("External Order Id", "ID AUTODOC"),
                ("SO number", "SO number"),
                ("Name", "Item Name"),
                ("Brand", "Brand"),
                ("SKU", "SKU"),
                ("Internal reference", "Internal reference"),
                ("Item description", "Item description"),
                ("Quantity", "Quantity"),
                ("GTIN (EAN)", "EAN"),
                ("Date expédition", None),
                ("Date réception", None),
            ]

            df_final = pd.DataFrame()
            for col_name, new_name in final_order:
                df_final[new_name or col_name] = df_f1[col_name] if col_name in df_f1.columns else ""

            # Mapping des palettes et remplissage des colonnes fixes
            df_final["N° PALETTE"] = df_final["N° COLIS"].map(colis_to_palette)
            df_final["Fournisseur"] = "MARKETPARTS"
            df_final["EAN"] = df_final["EAN"].astype(str).str.zfill(13)

            # Export vers Excel
            output = BytesIO()
            df_final.to_excel(output, index=False)
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active

            # Mise en page
            ws.insert_rows(1, amount=9)
            header_row_idx = 10
            headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[header_row_idx]]
            if "n° colis" not in headers:
                st.error("❌ Erreur : La colonne 'N° COLIS' n'a pas été trouvée dans la ligne d'en-tête (ligne 10).")
                return

            ws["I4"] = "Packing List/ Colisage"
            ws["I4"].font = Font(name="Helvetica", size=36)
            client_name = str(df_f2.iloc[0, 0])
            ws["E7"] = client_name
            unique_palettes = set(ws.cell(row=i, column=2).value for i in range(11, ws.max_row + 1) if ws.cell(row=i, column=2).value)
            ws["G7"] = len(unique_palettes)

            # Nettoyage du fond
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            for row in ws.iter_rows():
                for cell in row:
                    cell.fill = white_fill

            # Style de l'en-tête
            for col in range(1, 16):
                cell = ws.cell(row=10, column=col)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Bordures et alignement
            border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
            for row in ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=1, max_col=15):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustement automatique des colonnes
            for col in ws.columns:
                col_letter = col[0].column_letter
                if col_letter > "O":
                    continue
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = max_length + 2

            # Propriétés d'impression
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "10:10"
            ws.oddFooter.center.text = "Page &[Page] / &[Pages]"

            # Logo
            try:
                logo = OpenpyxlImage("logo_marketparts.png")
                logo.width = int(logo.width * 0.36)
                logo.height = int(logo.height * 0.36)
                ws.add_image(logo, "A1")
            except Exception as e:
                st.warning(f"⚠️ Erreur logo : {e}")

            # Vérification visuelle des correspondances
            try:
                def normalize(val):
                    return str(val).strip().replace('\xa0', '').replace('\n', '').replace('\r', '').lower()

                f1_series = df_f1["Document number"].dropna().astype(str).map(normalize)
                f1_series = f1_series[f1_series != 'nan']

                f2_series = df_f2["Package Number"].dropna().astype(str).map(normalize)
                f2_series = f2_series[f2_series != 'nan']

                f1_values = set(f1_series)
                f2_values = set(f2_series)

                only_in_f1 = sorted(f1_values - f2_values)
                only_in_f2 = sorted(f2_values - f1_values)

                if only_in_f1 or only_in_f2:
                    st.markdown("### ⚠️ Résumé des écarts entre F1 et F2")
                    st.markdown("---")

                if only_in_f1:
                    st.markdown(
                        f"""
                        <div style="background-color:#ff4d4d;padding:16px;border-radius:8px;color:white;">
                            <strong>❌ {len(only_in_f1)} document(s) trouvés dans F1 mais absents de F2</strong><br>
                            Exemples : {", ".join(only_in_f1[:10])}{'...' if len(only_in_f1) > 10 else ''}
                        </div>
                        """, unsafe_allow_html=True)

                if only_in_f2:
                    st.markdown(
                        f"""
                        <div style="background-color:#fff3cd;padding:16px;border-radius:8px;color:#856404;">
                            <strong>⚠️ {len(only_in_f2)} package(s) trouvés dans F2 mais absents de F1</strong><br>
                            Exemples : {", ".join(only_in_f2[:10])}{'...' if len(only_in_f2) > 10 else ''}
                        </div>
                        """, unsafe_allow_html=True)

            except Exception as e:
                st.error(f"❌ Erreur lors de la vérification des correspondances F1/F2 : {e}")

            # Export final
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            st.success("✅ Fichier généré avec succès")
            st.download_button(
                "📥 Télécharger le fichier formaté",
                data=final_output,
                file_name="PackingList_Formatée.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Erreur : {e}")

if __name__ == "__main__":
    run()

