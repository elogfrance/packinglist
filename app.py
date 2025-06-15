import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from copy import copy
from io import BytesIO
from PIL import Image

# Configuration de la page
st.set_page_config(
    page_title="Générateur de Packing List",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Affichage du logo
logo = Image.open("logo_marketparts.png")
st.image(logo, width=400)

# Titre principal
st.markdown(
    "<h1 style='color:#3a4e9f; font-size:24px;'>Générateur de Packing List</h1>",
    unsafe_allow_html=True
)

# Import F1 et F2 en colonnes séparées
col1, col2 = st.columns(2)

with col1:
    uploaded_f1 = st.file_uploader("📁 1. Importer le fichier TO SHIP", type=["xlsx"], key="f1")
    with st.expander("📎 Voir le template (F1)"):
        st.image("template_f1.png", caption="Exemple de fichier TO SHIP")

with col2:
    uploaded_f2 = st.file_uploader("📁 2. Importer le fichier E LOG", type=["xlsx"], key="f2")
    with st.expander("📎 Voir le template (F2)"):
        st.image("template_f2.png", caption="Exemple de fichier E LOG")

# Traitement à l'appui du bouton
if uploaded_f1 and uploaded_f2:
    try:
        # Chargement des fichiers
        wb_f1 = load_workbook(uploaded_f1)
        ws_f1 = wb_f1.active
        wb_f2 = load_workbook(uploaded_f2, data_only=True)
        ws_f2 = wb_f2.active

        # Suppression des colonnes G, H et I (colonnes 7, 8, 9)
        for col_idx in [9, 8, 7]:  # I, H, G
            ws_f1.delete_cols(col_idx)

        # Fusion de la cellule "Delivery Note / Bon de livraison" de A à H (attention : devient A à E après suppression)
        for row in ws_f1.iter_rows(min_row=1, max_row=20, max_col=8):
            for cell in row:
                if cell.value == "Delivery Note / Bon de livraison":
                    try:
                        ws_f1.unmerge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=8)
                    except:
                        pass
                    ws_f1.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=5)
                    break

        # Nettoyage de H9/I9 (devenues E9/F9 après suppression)
        h9 = ws_f1["E9"].value or ""
        try:
            i9 = ws_f1["F9"].value or ""
        except:
            i9 = ""
        ws_f1["E9"].value = f"{h9} {i9}".strip()

        try:
            ws_f1["F9"].value = None
        except:
            pass

        try:
            ws_f1.merge_cells("E9:F9")
        except:
            pass

        # Ajout du champ "N° de palette" en G11 (devenu colonne E après suppression)
        ws_f1["E11"].value = "N° de palette"
        if ws_f1["D11"].has_style:
            for attr in ["font", "border", "fill", "number_format", "protection", "alignment"]:
                setattr(ws_f1["E11"], attr, copy(getattr(ws_f1["D11"], attr)))

        # Recherche V depuis F2 vers F1 (sur colonne E maintenant)
        for row in range(12, ws_f1.max_row + 1):
            key = ws_f1[f"A{row}"].value
            if not key:
                continue
            for r in ws_f2.iter_rows(min_row=1, max_col=5):
                if r[3].value == key:
                    ws_f1[f"E{row}"].value = r[4].value
                    break

        # Sauvegarde dans la mémoire
        output = BytesIO()
        wb_f1.save(output)
        output.seek(0)

        # Bouton de téléchargement
        st.download_button(
            label="Télécharger packing list excel",
            data=output.getvalue(),
            file_name="PackingList.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Une erreur est survenue : {e}")
