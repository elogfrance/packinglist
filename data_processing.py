from copy import copy
from io import BytesIO
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


def load_excel_files(f1_file, f2_file) -> Tuple[Workbook, Worksheet, Worksheet]:
    """Load F1 and F2 Excel files and return workbook and worksheets."""
    wb_f1 = load_workbook(f1_file)
    ws_f1 = wb_f1.active
    wb_f2 = load_workbook(f2_file, data_only=True)
    ws_f2 = wb_f2.active
    return wb_f1, ws_f1, ws_f2


def clean_and_merge_columns(ws_f1: Worksheet) -> None:
    """Clean columns, merge cells and tidy header lines in F1."""
    headers = [cell.value for cell in ws_f1[11]]
    for idx in sorted(
        [i for i, h in enumerate(headers) if h in ["Unit Price", "Total Price"]],
        reverse=True,
    ):
        ws_f1.delete_cols(idx + 1)

    for row in ws_f1.iter_rows(min_row=1, max_row=20, max_col=8):
        for cell in row:
            if cell.value == "Delivery Note / Bon de livraison":
                try:
                    ws_f1.unmerge_cells(
                        start_row=cell.row,
                        start_column=1,
                        end_row=cell.row,
                        end_column=9,
                    )
                except Exception:
                    pass
                ws_f1.merge_cells(
                    start_row=cell.row, start_column=1, end_row=cell.row, end_column=8
                )
                break

    h9 = ws_f1["H9"].value or ""
    try:
        i9 = ws_f1["I9"].value or ""
    except Exception:
        i9 = ""
    ws_f1["H9"].value = f"{h9} {i9}".strip()

    try:
        ws_f1["I9"].value = None
    except Exception:
        pass

    try:
        ws_f1.merge_cells("H9:I9")
    except Exception:
        pass


def copy_cell_style(source, target) -> None:
    """Copy cell style from source to target cell."""
    if source.has_style:
        for attr in [
            "font",
            "border",
            "fill",
            "number_format",
            "protection",
            "alignment",
        ]:
            setattr(target, attr, copy(getattr(source, attr)))


def fill_palette_number(ws_f1: Worksheet, ws_f2: Worksheet) -> None:
    """Fill column H with palette numbers found in F2."""
    for row in range(12, ws_f1.max_row + 1):
        key = ws_f1[f"A{row}"].value
        if not key:
            continue
        for r in ws_f2.iter_rows(min_row=1, max_col=5):
            if r[3].value == key:
                ws_f1[f"H{row}"].value = r[4].value
                break


def save_workbook_to_bytes(wb: Workbook) -> BytesIO:
    """Return workbook content as BytesIO."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
